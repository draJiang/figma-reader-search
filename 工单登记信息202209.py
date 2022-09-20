# -*- coding:utf-8 -*-
from inspect import trace
import requests
import json
import os
import time
from openpyxl import load_workbook

from lxml import etree
from requests.models import cookiejar_from_dict

GCP_COOKIE = '_ntes_nnid=bcb44c991694f96db04f9a31683243c0,1644569150558; hb_MA-8D32-CBB074308F88_source=docs.popo.netease.com; hb_MA-9055-6BF1AAD03512_source=login.netease.com; visited_surveyid151848=166236583545376928; answered_surveyid151848=166236583545376928; answered_count_151848=1; answered_count_expire151848=2592000; mp_versions_hubble_jsSDK=DATracker.globals.1.6.14; HRBP-SESSIONID=Ghr-1663560550084-d417d201-93c9-4cbf-a011-8f8e5fb5bafe; hb_MA-ACDA-BCCED4A26C8A_source=www.figma.com; _km_noosfero_session=eyJwcm9maWxlIjoyNjI5MiwidXNlciI6Mjk1MTF9--b325f731ee3822ac40c660c9a41fe273c7117309; _my_redmine=BAh7EEkiCXVzZXIGOgZFRnsGSSIHNzkGOwBUaQJoBUkiD215X3VzZXJfaWQGOwBGaQK1f0kiDnVzZXJfbmFtZQY7AEZJIg9K5rGf5a2Q6b6ZBjsAVEkiDnVzZXJfbWFpbAY7AEZJIiFqaWFuZ3ppbG9uZ0Bjb3JwLm5ldGVhc2UuY29tBjsAVEkiCmN0aW1lBjsARmwrBzzQJ2NJIgphdGltZQY7AEZsKwfGVyhjSSIPc2Vzc2lvbl9pZAY7AFRJIiU0MGUyYTZlMTE0YjhlM2E0MzAzNDA3MjE3Y2RiNjI4YgY7AFRJIhBfY3NyZl90b2tlbgY7AEZJIjF3Skw1Y2lPM2U0TlQ3c0xqUEYyckhwQjhwQXdqNWNYSDRqSVNncGV2Tit3PQY7AEZJIh5sYXN0X3NlbGVjdGVkX2luc19wcm9qZWN0BjsARnsGSSIHNzkGOwBUaRxJIhZpc3N1ZXNfaW5kZXhfc29ydAY7AEZJIgxpZDpkZXNjBjsAVEkiDXBlcl9wYWdlBjsARmlp--9aaadc5ce531269929c4ba55867f35b4b79f6370'

def getGCPInfo(id):
    '''
    id:项目 id，例如 13303 / #13303
    platformL:易协作项目名称 ccmarketing/icc/cbg
    返回易协作工单信息
    '''
    # time.sleep(0.1)

    url = 'https://icc.pm.netease.com/api/v6/issue_show/base'

    headers = {'cookie': GCP_COOKIE,
               'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'}
    
    # id = str(id).replace('#','')
    if(id == None):
        print('id 异常')
        return 'id 异常'
    try:
        req = requests.post(url=url, headers=headers, json={'id': int(id)})
    except:
        print('erro')

    req = json.loads(req.text)

    if(req['res_code'] == 0):
        # 任务单不存在
        return req

    if(req['res_code'] == -1 or req['res_code'] == -2):
        # 无权限
        return req

    return req

    # # ID
    # pjID = req['data']['base']['id']
    # # 需求名称
    # subject = req['data']['base']['subject']
    # # 需求作者
    # author = req['data']['base']['author']['name']
    # # 需求描述
    # description = req['data']['base']['description']
    # # 跟踪标签
    # tracker = req['data']['base']['tracker']
    # # 期望交付时间
    # cf_fields = req['data']['cf_fields']
    # delayTime = ''
    # for item in cf_fields:
    #     if(item['name'] == '期望交付时间'):
    #         delayTime = item['value']

    # if(req['res_code'] == 1):
    #     return {'res_code': req['res_code'], 'delayTime': delayTime, 'id': pjID, 'subject': subject, 'author': author, 'description': description, 'tracker': tracker}

def getColumnIndex(clk_data_ws, name):
    '''
    获取指定表头的列索引
    clk_data_ws:目标表格， clk_data_wb.worksheets[0]
    name:目标表头名称，str
    '''
    clk_data_max_column = clk_data_ws.max_column  # 表格的最大列数

    for i in range(clk_data_max_column):
        i += 1  # i 不能取 0
        if(clk_data_ws.cell(row=1, column=i).value.find(name) >= 0):
            # ID 所在列
            return i


# 读取表格 ====================



# 表格地址
xlAdd = '/Users/cc/Downloads/任务列表的副本.xlsx'
# 表格
wb = load_workbook(xlAdd)
# 第一个 sheet
ws = wb.worksheets[0]  # 数据表，默认取首个 sheet

# 目标数据在第几列
ws_dh_index = getColumnIndex(ws, '#')
ws_ywbq_index = getColumnIndex(ws, '业务标签')
ws_xqbq_index = getColumnIndex(ws, '需求标签')

ws_max_row = ws.max_row

# 遍历数据表
for index in range(ws_max_row):

    index+=1

    if(index==1):
        # 跳过表头
        continue

    # 进度条
    if(index == ws_max_row):
        print('\r 完成', end='')
    else:
        print('\r'+str(index)+'/'+str(ws_max_row), end='')


    
    dh = ws.cell(row=index, column=ws_dh_index).value
    
    # 获取当前工单信息
    req = getGCPInfo(dh)
    
    # 获取总单信息
    try:
        if(len(req['data']['base']['ancestors'])==0):
            zd_req = req
        else:
            zd_req = getGCPInfo(str(req['data']['base']['ancestors'][0]['id']))
    
        # 业务标签
        ywbq = zd_req['data']['cf_fields'][6]['value']
        subject = zd_req['data']['base']['subject']
        # 需求标签
        
        # 将易协作数据写入表格中 ================
        ws.cell(row=index, column=ws_ywbq_index).value = ywbq
        ws.cell(row=index, column=ws_xqbq_index).value = subject

    except:
        continue

# 另存为表格 ================
wb.save(xlAdd)